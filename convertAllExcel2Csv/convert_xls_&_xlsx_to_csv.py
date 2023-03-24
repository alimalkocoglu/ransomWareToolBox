import csv
import pyexcel as xsl_conv
from openpyxl import load_workbook
import glob
import re
from time import perf_counter
import datetime
import logging
# By default openpyxl does not guard against quadratic blowup or billion laughs xml attacks. To guard against these attacks install defusedxml.


def xls_to_xlsx (current_filename) -> None:
  ''' receives a file xls file name from the same directory converts it to .xlsx file'''

  new_file_name = re.sub('.xls','.xlsx',current_filename)
  xsl_conv.save_book_as(file_name=current_filename, dest_file_name=new_file_name)
  return


# converts all xsl files in the dir

extension = '.xlsx'


# ask user input for sheets


def allSheets(filename, all_Sheets=True):
    ''' receives an file name from the same directory and calls the convert2Csv function for all the sheets.'''

    wb = load_workbook(filename=filename, read_only=True)
    workBookName = re.sub('\.xlsx','',filename)

    if all_Sheets:
        sheets = wb.sheetnames[:]
    else:
        sheets = wb.sheetnames[:1]

    for sheetName in sheets:
        csv_file = f"{filename.strip()}_{sheetName}.csv"
        sheet = wb[sheetName]
        with open (f"{workBookName.strip()}_{sheetName}.csv",'w') as csv_output:
            writer = csv.writer(csv_output, delimiter=',')

            for line in sheet.iter_rows(values_only=True):
                writer.writerow(line)


if __name__ == '__main__':
    log = logging.getLogger()
    logging.basicConfig(filename=f"{datetime.date.today()}.log", level="DEBUG")

    INPUT = "input"
    OUTPUT = "output"

    print("* "*20)
    print("\nThe script will convert all the xlsx files in the current directory.\n")
    print("* "*20)
    print("\n")
    print("To convert only the first sheet type 1 to convert all the sheets in a workbook type 2 and hit enter")
    user_input = input("your choice:")
    all_Sheets = False
    if user_input == "2":
        all_Sheets = True

    xsl_file_list = glob.glob('*.xls')

    for file in xsl_file_list:
        logging.info(f"Processing {file}")
        start = perf_counter()
        try:
            xls_to_xlsx(file)
        except Exception as e:
            logging.error(e)
        logging.info(f"Processing took {perf_counter()- start:2f}")

    print("all sheets wanted")
    for file in glob.glob(f"*{extension}"):
        start = perf_counter()
        logging.info(f"Processing {file}")
        try:
            allSheets(file, all_Sheets)
        except Exception as e:
            logging.error(e)
        logging.info(f"Processing took {perf_counter()- start:2f}")
